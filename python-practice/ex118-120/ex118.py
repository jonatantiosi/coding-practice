'''
118 — Ler planilha Excel e modificar célula
Tarefa: abra uma planilha .xlsx, percorra as linhas a partir da segunda linha
e, ao encontrar um nome específico, altere o valor de outra célula da mesma
linha.
Conceitos: openpyxl.load_workbook, Worksheet.iter_rows, Worksheet.cell, save.
'''
from pathlib import Path
from openpyxl import load_workbook, Workbook
from typing import cast

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = cast(Workbook, load_workbook(WORKBOOK_PATH))
worksheet = workbook['Sheet']

for row in worksheet.iter_rows(min_row=2):
    for cell_ in row:
        if cell_.value == 'Alberto':
            worksheet.cell(cell_.row, 3, 9)

workbook.save(WORKBOOK_PATH)
