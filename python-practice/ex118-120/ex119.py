'''
119 — Criar planilha Excel com dados
Tarefa: crie um arquivo Excel do zero, adicione cabeçalhos e insira dados de
uma lista usando append.
Conceitos: openpyxl.Workbook, workbook.active, Worksheet.append, save.
'''
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import cast

data = [
    ['Jonatan', 24, 1.8],
    ['Paulo', 24, 1.81],
    ['Maple', 6, 0.2],
]

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook_new.xlsx'

workbook = Workbook()
worksheet = cast(Worksheet, workbook.active)
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Altura')

for item in data:
    worksheet.append(item)

workbook.save(WORKBOOK_PATH)
