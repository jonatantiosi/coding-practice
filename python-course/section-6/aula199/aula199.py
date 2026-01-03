# mypy: ignore-errors
# flake8: noqa

''' reading '''

# openpyxl - ler e alterar dados de uma planilha
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.cell import MergedCell
from typing import cast


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook_2.xlsx'

# carregando arquivo do xlsx
workbook = cast(Workbook, load_workbook(WORKBOOK_PATH))
# nomear planilha
sheet_name = 'Planilha1'
# selecionar a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell | MergedCell, ...]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')
        if cell.value == 'Luiz':
            worksheet.cell(cell.row, 2, 65)
    print()

# valor da celula B#
cell_b3 = cast(Cell, worksheet['B3'])
print(cell_b3.value)
cell_b3.value = 20

workbook.save(WORKBOOK_PATH)


''' creating_1 '''

# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheet = cast(Worksheet, workbook.active)   # planilha ativa
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')


students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i in range(2, 10):
#     for j in range(1, 4):
#         print('Linha', i, 'Coluna', j)

# logica 1, mais dificil mas pode ser util futuramente
# for i, student_row in enumerate(students, start=2):
#     # print(i, student_row)
#     for j, student_column in enumerate(student_row, start=1):
#         # print(i, j, student_column)
#         worksheet.cell(i, j, cast(str, student_column))

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)

print(workbook.sheetnames)

''' creating_2 '''

# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
# from typing import cast


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook_2.xlsx'

workbook = Workbook()

# nome da planilha
sheet_name = 'Planilha1'
# criacao da planilha
workbook.create_sheet(sheet_name, 0)  # 0 garante que seja a primeira na ordem
# selecionar a planilha
worksheet: Worksheet = workbook[sheet_name]
# remover planilha padrao "sheets"
workbook.remove(workbook['Sheet'])

# criacao dos cabecalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')


students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i in range(2, 10):
#     for j in range(1, 4):
#         print('Linha', i, 'Coluna', j)

# logica 1, mais dificil mas pode ser util futuramente
# for i, student_row in enumerate(students, start=2):
#     # print(i, student_row)
#     for j, student_column in enumerate(student_row, start=1):
#         # print(i, j, student_column)
#         worksheet.cell(i, j, cast(str, student_column))

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)

