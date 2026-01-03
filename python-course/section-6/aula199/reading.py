# openpyxl - ler e alterar dados de uma planilha
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.cell import MergedCell
from typing import cast


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook_2.xlsx'

# carregando arquivo do xlsx
workbook = cast(Workbook, load_workbook(WORKBOOK_PATH))
# nomear planilha
sheet_name = 'Planilha1'
# selecionar a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell | MergedCell, ...]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')
        if cell.value == 'Luiz':
            worksheet.cell(cell.row, 2, 65)
    print()

# valor da celula B#
cell_b3 = cast(Cell, worksheet['B3'])
print(cell_b3.value)
cell_b3.value = 20

workbook.save(WORKBOOK_PATH)
